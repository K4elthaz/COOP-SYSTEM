import openpyxl
import pandas as pandas
import os


def load_workbook(ex_path):
    if os.path.exists(ex_path):
        return openpyxl.load_workbook(ex_path)
    else:
        return "File not Found...!"
    
ex_path = "Member.xlsx"
ex = load_workbook(ex_path)
sheet = ex["Sheet"]
sheet_obj = ex.active

